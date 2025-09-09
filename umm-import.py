import csv
import logging
import openpyxl

import common


def amend_klassenkuerzel_and_bewerbskuerzel(ws):
    logging.debug("amend klassenkuerzel and bewerbskuerzel")
    cell = ws['K1']
    cell.value = 'Klassenkürzel'
    cell = ws['L1']
    cell.value = 'Bewerbskürzel'
    for row_index, cell in enumerate(ws['G']):
        if row_index == 0:
            assert cell.value == 'Kategorie'
        elif cell.value is not None:
            name = []
            column_row_code = 'D{}'.format(cell.row)
            name_cell = ws[column_row_code]
            name.append(name_cell.value)
            column_row_code = 'C{}'.format(cell.row)
            name_cell = ws[column_row_code]
            name.append(name_cell.value)
            name = ' '.join(name)
            logging.info('name: %s' % name)
            kategorie = cell.value.upper()
            klassenkuerzel = common.get_klassenkuerzel(kategorie)
            bewerbskuerzel = common.get_bewerbskuerzel(kategorie)
            logging.info('%s %s: %r => %r, %r' % (row_index, name, kategorie, klassenkuerzel, bewerbskuerzel))
            column_row_code = 'K{}'.format(cell.row)
            target_cell = ws[column_row_code]
            target_cell.value = klassenkuerzel
            column_row_code = 'L{}'.format(cell.row)
            target_cell = ws[column_row_code]
            target_cell.value = bewerbskuerzel


def write_csv(ws, filename):
    logging.debug('writing csv')
    with open(filename, 'w', newline="", encoding='latin-1') as file_handle:
        csv_writer = csv.writer(file_handle, delimiter=';', quotechar='"')
        for row in ws.iter_rows():
            csv_writer.writerow([cell.value for cell in row])


def main(args):
    common.setup_logging(args.verbose)
    logging.debug("main(%s)" % args)
    wb = openpyxl.load_workbook(args.subscriptions)
    ws = wb.active
    amend_klassenkuerzel_and_bewerbskuerzel(ws)
    csv_filename = args.subscriptions.rstrip('xlsx') + 'csv'
    write_csv(ws, csv_filename)
    logging.debug("done")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='UMM Import Tool')
    parser.add_argument('-v', '--verbose', action="store_true", help="be verbose")
    parser.add_argument('subscriptions', help='XLSX file')
    args = parser.parse_args()

    main(args)
